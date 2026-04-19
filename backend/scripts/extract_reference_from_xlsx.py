"""Extract reference data (countries + variables) from the Excel prototype into seed CSVs.

Re-run whenever the prototype is updated. Outputs:
  supabase/seeds/countries.csv — iso3, name, region, as_of_year, hdi_value, segment
  supabase/seeds/variables.csv — code, name, category, direction, is_quantitative, description
"""
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX = REPO_ROOT / "prototype" / "Country Prototype Original HDI with Ridge.xlsx"
SEEDS_DIR = REPO_ROOT / "supabase" / "seeds"

REGION_MAP = {
    "AFRICA": "AFRICA",
    "EMERGING MARKET NON-AFRICA": "EMERGING_NON_AFRICA",
    "DEVELOPED MARKET": "DEVELOPED",
}

HDI_TO_VALUE = {
    "VERY HIGH HUMAN DEVELOPMENT": 0.9,
    "HIGH HUMAN DEVELOPMENT": 0.75,
    "MEDIUM HUMAN DEVELOPMENT": 0.6,
    "LOW HUMAN DEVELOPMENT": 0.45,
}

# Built from the Excel "High Segment" driver list (column 'Low mapping' at col I)
# plus the "Variable Definitions" sheet. See spec §4.2.
VARIABLES = [
    # code,                 name,                                 category,         direction,        is_quant, description
    ("dcpi_5_adj",          "Inflation 5 year average",           "Economic",       "higher_worse",   True,     "Average CPI % change over 5 years."),
    ("nom_rir_vol",         "Nominal Interest Rate Volatility",   "Economic",       "higher_worse",   True,     "Volatility of nominal interest rates."),
    ("gdp_capita",          "GDP per Capita",                     "Economic",       "higher_better",  True,     "GDP per head (US$)."),
    ("growth_vol",          "Growth Volatility",                  "Economic",       "higher_worse",   True,     "Stdev of Real GDP % change over 5 years."),
    ("macro_var",           "Macroeconomic",                      "Economic",       "higher_better",  False,    "Qualitative macroeconomic stability score."),
    ("atf",                 "Access to Finance",                  "Currency",       "higher_better",  False,    "Qualitative access-to-finance score."),
    ("dt",                  "Debt Trend",                         "Currency",       "higher_worse",   True,     "3-year trend in foreign debt to GDP."),
    ("fdg_3yr",             "Foreign Debt to GDP 3 year average", "Currency",       "higher_worse",   True,     "Average foreign debt / GDP over 3 years."),
    ("cof",                 "Cost of Funds",                      "Finance",        "higher_worse",   True,     "Money market interest rate (%)."),
    ("pr",                  "Political Risk Environment",         "Political",      "higher_better",  False,    "Qualitative political risk score."),
    ("rol",                 "Rule of Law",                        "Business Risk",  "higher_better",  False,    "World Governance Indicator: Rule of Law."),
    ("db",                  "Doing Business",                     "Business Risk",  "higher_better",  False,    "World Bank Doing Business qualitative score."),
    ("sr",                  "Security Risk",                      "Risk",           "higher_better",  False,    "Global Insights security risk score."),
    ("debt_service_ratio",  "Debt Service Ratio 1 year",          "Currency",       "higher_worse",   True,     "Debt service due as % of exports + remittances."),
]


def main() -> None:
    SEEDS_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX, data_only=True)

    # --- countries ---
    country_sheet = wb["Country List"]
    # Work out which analytical segment each country belongs to by scanning the three columns
    # (LOW, HIGH, No Data) under their headers. Those are sorted name lists.
    rows = list(country_sheet.iter_rows(min_row=2, values_only=True))

    low_names = {r[5] for r in rows if r[5]}
    high_names = {r[6] for r in rows if r[6]}
    nodata_names = {r[7] for r in rows if r[7]}

    country_rows = []
    for r in rows:
        name, iso3, hdi_label, em_dm = r[0], r[1], r[2], r[3]
        if not (name and iso3):
            continue
        if name in high_names:
            segment = "HIGH"
        elif name in low_names:
            segment = "LOW"
        elif name in nodata_names:
            segment = "NODATA"
        else:
            # Fall back by HDI band: MEDIUM → LOW per the prototype's grouping
            segment = "NODATA"
        country_rows.append({
            "iso3": iso3.strip(),
            "name": name.strip(),
            "region": REGION_MAP.get((em_dm or "").strip(), "UNKNOWN"),
            "as_of_year": 2011,  # prototype baseline year per Variable Definitions sheet
            "hdi_value": HDI_TO_VALUE.get((hdi_label or "").strip(), None),
            "segment": segment,
        })

    countries_csv = SEEDS_DIR / "countries.csv"
    with countries_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["iso3", "name", "region", "as_of_year", "hdi_value", "segment"])
        writer.writeheader()
        writer.writerows(country_rows)
    print(f"Wrote {len(country_rows)} countries -> {countries_csv}")

    # --- variables ---
    variables_csv = SEEDS_DIR / "variables.csv"
    with variables_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["code", "name", "category", "direction", "is_quantitative", "description"])
        for v in VARIABLES:
            writer.writerow(list(v))
    print(f"Wrote {len(VARIABLES)} variables -> {variables_csv}")


if __name__ == "__main__":
    main()
