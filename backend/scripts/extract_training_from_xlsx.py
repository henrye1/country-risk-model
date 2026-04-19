"""Extract 2011 training data for each HDI segment from the Excel prototype.

Outputs two CSVs in `supabase/seeds/`, one per segment. Each row is a country-year
observation used to train a Ridge regression.

Columns (in order):
  iso3, name, segment, eiu_score,
  dcpi_5_adj, nom_rir_vol, gdp_capita, growth_vol, macro_var,
  atf, dt, fdg_3yr, cof,
  pr, rol, db, sr, debt_service_ratio

The 14 driver columns match the `variables.code` seeded in Plan 1. `eiu_score` is
the target variable (existing country rating). Missing values are left blank — the
training code in Task 7 handles NaN row-dropping.
"""
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX = REPO_ROOT / "prototype" / "Country Prototype Original HDI with Ridge.xlsx"
SEEDS_DIR = REPO_ROOT / "supabase" / "seeds"

# Map our variable codes to the raw-sheet column headers.
# Each value is a list of candidate header names to search for (first match wins).
DRIVER_MAP: dict[str, list[str]] = {
    "dcpi_5_adj":          ["dcpi_2007_2011"],       # 5-year average inflation
    "nom_rir_vol":         ["nom_rir_vol_2011"],
    "gdp_capita":          ["gdp_capita_2011"],
    "growth_vol":          ["growth_vol_2010_2011"],
    "macro_var":           ["ic_2011"],              # "investment climate" used as macro qualitative proxy
    "atf":                 ["atf_2011"],
    "dt":                  ["dt_2011"],
    "fdg_3yr":             ["fdg_2011"],             # approximation: use 2011 level (3yr avg not in raw)
    "cof":                 ["cof_2011"],
    "pr":                  ["pr_2011"],
    "rol":                 ["rol_2011"],
    "db":                  ["db_2011"],
    "sr":                  ["sr_2011"],
    "debt_service_ratio":  ["dsr_2011"],
}

# Target column (EIU rating) — search candidates across common Excel naming styles.
EIU_CANDIDATES = ["eiu_2011", "EIU_score", "EIU_2011", "eiu_score", "EIU"]


def _header_index(header: list[str | None], names: list[str]) -> int | None:
    lower = [str(h).strip().lower() if h is not None else "" for h in header]
    for n in names:
        needle = n.strip().lower()
        if needle in lower:
            return lower.index(needle)
    return None


def _extract_one_sheet(wb, sheet_name: str, segment: str, out_csv: Path) -> int:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])

    iso_idx = _header_index(header, ["iso_code"])
    name_idx = _header_index(header, ["country"])
    eiu_idx = _header_index(header, EIU_CANDIDATES)

    driver_indices: dict[str, int | None] = {
        code: _header_index(header, cands) for code, cands in DRIVER_MAP.items()
    }

    out_rows = []
    for r in rows[1:]:
        iso = r[iso_idx] if iso_idx is not None else None
        name = r[name_idx] if name_idx is not None else None
        if not (iso and name):
            continue
        row: dict[str, object] = {
            "iso3": str(iso).strip(),
            "name": str(name).strip(),
            "segment": segment,
            "eiu_score": r[eiu_idx] if eiu_idx is not None else None,
        }
        for code, idx in driver_indices.items():
            row[code] = r[idx] if idx is not None else None
        out_rows.append(row)

    fieldnames = ["iso3", "name", "segment", "eiu_score"] + list(DRIVER_MAP.keys())
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in out_rows:
            w.writerow({k: ("" if row[k] is None else row[k]) for k in fieldnames})
    return len(out_rows)


def main() -> None:
    SEEDS_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX, data_only=True)

    n_high = _extract_one_sheet(wb, "High_Raw", "HIGH", SEEDS_DIR / "training_high.csv")
    n_low = _extract_one_sheet(wb, "Low_Raw", "LOW", SEEDS_DIR / "training_low.csv")

    print(f"Wrote {n_high} HIGH training rows -> {SEEDS_DIR / 'training_high.csv'}")
    print(f"Wrote {n_low} LOW training rows -> {SEEDS_DIR / 'training_low.csv'}")


if __name__ == "__main__":
    main()
